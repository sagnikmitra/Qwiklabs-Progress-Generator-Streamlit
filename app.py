import streamlit as st
import numpy as np
import pandas as pd
import sklearn
from sklearn import datasets
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt


gc = pd.read_excel('gc.xlsx')
import openpyxl as pxl
from openpyxl import load_workbook

exp = pxl.load_workbook('gc.xlsx')
sheet = exp['Sister Nivedita University, Kol']
st.write("""
# Check your Qwiklabs Progress
""")
flag = 0
str = st.text_input('Enter You Qwiklabs Email Id')
for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=2).value == str:
        st.write(sheet.cell(row=i, column=5).value)
        st.write(sheet.cell(row=i, column=6).value)
        st.write(int(sheet.cell(row=i, column=7).value))
        st.write(int(sheet.cell(row=i, column=8).value))
        flag = flag + 1

if flag==0 :
    st.write("No Search Found")


dataset_name = st.sidebar.selectbox("Select Dataset",("Iris","Breast Cancer","Wine Dataset"))
st.write(f"## Name of the Dataset: {dataset_name}")
classifier_name = st.sidebar.selectbox("Select Classifier",("KNN","SVM","Random Forest"))

# def get_quests:
#     pass

def get_dataset(dataset_name):
    if dataset_name == "Iris":
        data = datasets.load_iris()
    elif dataset_name == "Breast Cancer":
        data = datasets.load_breast_cancer()
    else:
        data = datasets.load_wine()
    X = data.data
    y = data.target
    return X,y

X, y = get_dataset(dataset_name)
st.write("## Shape of Dataset:",X.shape)
st.write("## Number of classes:",len(np.unique(y)))

def add_parameter_ui(classifier_name):
    params = dict()
    if classifier_name == "KNN":
        K = st.sidebar.slider("K",1,15)
        params["K"] = K
    elif classifier_name == "SVM":
        C = st.sidebar.slider("C",0.01,10.0)
        params["C"] = C
    elif classifier_name == "Random Forest":
        max_depth = st.sidebar.slider("max_depth",2,15)
        n_estimators = st.sidebar.slider("n_estimators",1,100)
        params["max_depth"] = max_depth
        params["n_estimators"] = n_estimators
    return params

params = add_parameter_ui(classifier_name)

def get_classifier(classifier_name, params):
    if classifier_name == "KNN":
        classifier = KNeighborsClassifier(n_neighbors=params["K"])
    elif classifier_name == "SVM":
        classifier = SVC(C=params["C"])
    elif classifier_name == "Random Forest":
        classifier = RandomForestClassifier(n_estimators=params["n_estimators"],max_depth=params["max_depth"],random_state=1234)
    return classifier

classifier = get_classifier(classifier_name,params)

X_train, X_test, y_train, y_test = train_test_split(X,y,test_size = 0.2, random_state = 1234)

classifier.fit(X_train,y_train)
y_pred = classifier.predict(X_test)

accuracy_score = accuracy_score(y_test, y_pred)

st.write(f"## Classifier = {classifier_name}")
st.write(f"## Accuracy = {accuracy_score}")

pca = PCA(2)
X_projected = pca.fit_transform(X)

x1 = X_projected[:,0]
x2 = X_projected[:,1]

fig = plt.figure()
plt.scatter(x1,x2,c=y,alpha=0.8,cmap = "viridis")
plt.xlabel("Principal Component 1")
plt.ylabel("Principal Component 2")
plt.colorbar()

st.pyplot(fig)
